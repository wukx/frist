# -*- coding:utf-8 -*-
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from .models import ExerContent
from openpyxl.compat import range


def import_user(self, request, obj, change):

    wb = load_workbook(filename=obj.file.path)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    rows = ws.max_row  # 获取读取的excel的文件的行数
    headers = ['No', 'title', 'option', 'answer', 'style']
    lists = []
    #users = request.user
    for row in range(2, rows):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        # for header in headers:
        title = cell['title']
        option = cell['option']
        answer = cell['answer']
        style = cell['style']
        sql = ExerContent(title=title, option=option, style=style, answer=answer)
        sqllist.append(sql)
    ExerContent.objects.bulk_create(sqllist)